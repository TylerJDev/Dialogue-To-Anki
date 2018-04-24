import unittest
import csv
import openpyxl
import os

def cleanupFiles(files=[]):
	for checkFiles in files:
		f_name, ext = os.path.splitext(checkFiles)
		if ext not in ['.txt', '.xlsx']:
			break;
		else:
			if os.path.exists(checkFiles):
				os.remove(checkFiles);

	for i in files:
		if os.path.exists(i):
			return True;
			
	return False;

def convertFromXLSX():
	global files
	wb = openpyxl.load_workbook('dialogueExport_Test.xlsx') # Gets the excel file, (Note that this is the same directory as this script)
	#sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	sheet = wb[wb.sheetnames[0]]
	
	# Check if sheet cell carries the proper name (For validation reasons)
	response_text = 'ONE';
	if sheet.cell(row=1, column=1).value != response_text:
		# Look for the column that carries the specified name
		for i in range (1, sheet.max_column): # To get audio filenames
			if sheet.cell(row=1, column=i).value == response_text:
				print('FOUND AT: ' + str(i));
	
	sheet_data = [];
	for i in range (1, sheet.max_row + 1):
		sheet_data.append(sheet.cell(row=i, column=1).value)
		
	with open('Exported_Test.txt', 'w+', encoding='utf-8') as exp:
		for i in sheet_data:
			exp.write(i + '\n');
		
	if os.path.exists('Exported_Test.txt'):
		cleanupFiles(['Exported_Test.txt', 'dialogueExport_Test.xlsx']);
		return True;


def convert2():
	global files
	# Possibly add tempfile, or cleanup to not make unnecessary clutter 
	d_export = [
	['ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'ONE'],
	['ONE_2', 2, 3, 4, 5],
	['ONE_3', 2, 3, 4, 5]
	]

	# Create the txt file
	if not os.path.exists('dialogueExport_TEST.txt'): 
		with open('dialogueExport_TEST.txt', "w", encoding='utf-8') as txt: # Creates the ini file with the selected language
			for i in d_export:
				for x in i:
					txt.write(str(x) + '\t');
				txt.write('\n');
	
	input_file = 'dialogueExport_TEST.txt'
	output_file = 'dialogueExport_TEST.xlsx'

	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]

	with open(input_file, 'r', encoding='utf-8') as data:
		reader = csv.reader(data, delimiter='\t')
		for row in reader:
			ws.append(row)

	wb.save(output_file)
	cleanupFiles([input_file])
	return True;

class TestGrabbing(unittest.TestCase):
	def testConvertingTxt2xlsx(self):
		self.assertTrue(convert2(), os.path.exists('dialogueExport_Test.xlsx'));
		
	def testXLSX2formats(self):
		self.assertTrue(convertFromXLSX());
		
if __name__ == '__main__':
    unittest.main()