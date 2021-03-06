import argparse
import os
import openpyxl
import csv

m_descr = ''' B_Languages is a program to get dialogue from BETHESDA game(s), and convert them into different
formats, (i.e, flashcards, csv, txt, etc). The following games are currently supported:
	- Skyrim
	- Skyrim SE
	- Fallout 4
	- Fallout NV
	
	Games that /may/ work:
		- Fallout 3
		- Oblivion
	
	To see support details, bugs and more, please read the README file.
'''

# Ask user for path to audio files, perferably in a 

sheet_data = [];
csv_data = [];
xlsx_data = [];
row_count = [];
audio_files = [];
supportedFormats = ['txt', 'csv']

def convertFromCSV(inputFile, game, language):
	global row_count;
	global csv_data
	global audio_files
	audio_files = [];
	csvData = {};
	wb = openpyxl.load_workbook(inputFile)
	sheet = wb[wb.sheetnames[0]]
	ignore_types = ['GenericDialgoueWounded', 'DialogueGiant', 'VoicePowers']
	responseGames = {
		'Skyrim': {'RESPONSE TEXT': '', 'QUEST': '', 'FULL PATH': '', 'FILENAME': ''}
	}
	
	def getColumnsRows():
		for getColumns in range(1, sheet.max_column):
			if sheet.cell(row=1, column=getColumns).value == 'FULLPATH':
					responseGames[game]['FULL PATH'] = getColumns;
			else:
				if sheet.cell(row=1, column=getColumns).value in responseGames[game]:
					responseGames[game][sheet.cell(row=1, column=getColumns).value] = getColumns;
		
	getColumnsRows()
	abr = language[:2].upper()
	# Loop through audio filenames
	for getAudioNames in range(1, sheet.max_row):
		skip = 0;
		try:
			voice_types = list(filter(lambda plc_var: plc_var in sheet.cell(row=getAudioNames, column=responseGames[game]['QUEST']).value, ignore_types));
			skip = len(voice_types)
		except TypeError:
			pass
			
		if skip == 0: # Only skip if current row is one of the ignored types
			csvData[sheet.cell(row=getAudioNames, column=responseGames['Skyrim']['FULL PATH']).value] = '[sound:' + sheet.cell(row=getAudioNames, column=responseGames[game]['FILENAME']).value.lower() + '_' + abr + '.wav] ' + sheet.cell(row=getAudioNames, column=responseGames[game]['RESPONSE TEXT']).value
			audio_files.append(sheet.cell(row=getAudioNames, column=responseGames[game]['FILENAME']).value.lower() + '.fuz');
			
		
	csv_data.append(csvData);
def convertFromXLSX(inputFile, game, count, language):
	global sheet_data;
	global row_count;
	xlsxData = {};
	wb = openpyxl.load_workbook(inputFile)
	sheet = wb[wb.sheetnames[0]]
	ignore_types = ['GenericDialgoueWounded', 'DialogueGiant', 'VoicePowers']
	# Check if sheet cell carries the proper name (For validation reasons)
	
	# Response Text = The main dialogue, Quest = Types to ignore (For Skyrim this is the 'Quest' column, Full Path = Path to audio, Filename = name of audio file
	responseGames = {
		'Skyrim': {'RESPONSE TEXT': '', 'QUEST': '', 'FULL PATH': '', 'FILENAME': ''},
		'Fallout New Vegas': {'RESPONSE TEXT': '', 'QUEST': '', 'FULL PATH': '', 'FILENAME': ''},
		'Fallout 4': {'RESPONSE TEXT': '', 'QUEST': '', 'FULL PATH': '', 'FILENAME': ''}
	}

	abr = language[:2].upper()
	
	for getColumns in range(1, sheet.max_column):
		if sheet.cell(row=1, column=getColumns).value in responseGames[game]:
			responseGames[game][sheet.cell(row=1, column=getColumns).value] = getColumns;

	for i in range (1, sheet.max_row + 1):
		try:
			skip = 0;
			try:
				voice_types = list(filter(lambda plc_var: plc_var in sheet.cell(row=i, column=responseGames[game]['QUEST']).value, ignore_types));
				skip = len(voice_types)
			except TypeError:
				pass
				
			#if len(list(filter(lambda plc_var: plc_var in sheet.cell(row=i, column=3).value, ignore_types))) == 0:
				# Checks if NPC is Male or Lady
				# if 'female' in sheet.cell(row=i, column=6).value.lower():
					# print('FEMALE');
				# elif 'male' in sheet.cell(row=i, column=6).value.lower():
					# print('MALE');

			try:
				xlsxData[sheet.cell(row=i, column=responseGames[game]['FULL PATH']).value] = '[sound:' + sheet.cell(row=i, column=responseGames[game]['FILENAME']).value.lower() + '_' + abr + '.wav] ' + sheet.cell(row=i, column=responseGames[game]['RESPONSE TEXT']).value
			except TypeError:
				pass
				
			try:
				audio_files.append(sheet.cell(row=i, column=responseGames[game]['FILENAME']).value); # Append the audio filenames from file
			except TypeError:
				pass
		except (AttributeError, IndexError):
			pass;
	
	xlsx_data.append(xlsxData);
def convert2(inputFile, game, encoding=0):
	rn = game.gamePath()
	filename, fileExt = os.path.splitext(inputFile)
	input_file = inputFile
	# Example: 'Skyrim_DialogueExport_Spanish_0'
	output_file = game.gamePath() + '_DialogueExport_' + game.language + '.xlsx'
	
	if os.path.exists('dialogues/' + output_file): # If path exists already, will rename with prefix _int
		count = 0;
		choice = input('File already exists! Overwrite %s? (Y/N)' % output_file);
		if choice.lower() == 'y':
			def rename(c, output_file):
				if os.path.exists('dialogues/' + output_file + '_' + str(c)):
					c += 1
					rename(c);
				else:
					output_file = output_file + str(c)
			rename(count, output_file);
		else:
			return 'dialogues/' + output_file;
			
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	encoding_type = 'UTF-8';
	
	if fileExt == '.xlsx' or encoding != 0:
		encoding_type = 'ISO-8859-1';

	try:
		with open(input_file, 'r', encoding=encoding_type) as data:
			reader = csv.reader(data, delimiter='\t')
			print('Writing to file...');
			def appendRows(c=0, error_count=0):
				errorCount = error_count;
				count = c;
				try:	
					for row in reader:
						ws.append(row)
						count += 1;
					wb.save('dialogues/' + output_file)
					return 'dialogues/' + output_file;
				except Exception as e:
					errorCount += 1;
					if errorCount > 10:
						return False;
					return appendRows(count, errorCount);
				
			res = appendRows();
	except FileNotFoundError:
		# Check dialogues folder if conditions are met
		if os.path.exists('dialogues/' + input_file):
			return convert2('dialogues/' + input_file, game);
		else:
			print('Couldnt find file at ' + input_file + '!');
			return False;
	if res != False:
		return 'dialogues/' + output_file;
	else:
		return convert2(inputFile, game, encoding='ISO-8859-1')
def support(print_games=False, game='', language=''):
	# * Chinese = Traditional Chinese
	supportedGameDict = {'Fallout 4': [['fallout 4', 'f4', 'fallout4'], # Words that will translate to the games offical title (i.e, -g f4 > Fallout 4) 
	{'English': {'audio': True}, 'French': {'audio': True}, 'Italian': {'audio': True}, 'German': {'audio': True}, 'Spanish': {'audio': True}, 'Polish': {'audio': False}, 'Portuguese-Brazil': {'audio': False}, 'Russian': {'audio': False}, 'Chinese': {'audio': False}, 'Japanese': {'audio': True}}
	], 
	'Skyrim': [['skyrim', 'skyrim se', 'skyrim special edition', 'sky'], 
	{'English': {'audio': True}, 'French': {'audio': True}, 'Italian': {'audio': True}, 'German': {'audio': True}, 'Spanish': {'audio': True}, 'Polish': {'audio': True}, 'Chinese':
	{'audio': False, 'support': 'skyrim_se'}, 'Russian': {'audio': True}, 'Japanese': {'audio': True}, 'Czech': {'audio': True, 'support': 'skyrim'}}
	],
	'Fallout New Vegas': [['fnv', 'fallout new vegas', 'fallout nv'], 
	{'English': {'audio': True}, 'French': {'audio': True}, 'Italian': {'audio': True}, 'German': {'audio': True}, 'Spanish': {'audio': True}}
	]}
	
	s_arr = [];
	for i in supportedGameDict:
		if print_games == True:
			print(i + ': ' + ', '.join(supportedGameDict[i][0]) + '\n')
		s_arr.append(supportedGameDict[i][0]);
		
		if len(game) > 0 and game in supportedGameDict[i][0]:
			if len(language) > 0:
				check = [];
				# Loop through language list and check if both languages are supported, 1 = True, 0 = False
				for checkLang in language:
					if checkLang in supportedGameDict[i][1]:
						check.append(1);
					else: 
						check.append(0)
				if sum(check) == 2:
					return True;
				else:
					return language[check.index(0)]; # Returns the non-supported language
				
			return i;
			
		
	return sum(s_arr, []); # Combines supportGameDict arrays
	
# File location, Creation Kit language,
class games:
	## Store current game, that game folders location, and the language
	def __init__(self, game='', gameFolder='', language=''):
		self.game = game
		self.gameFolder = gameFolder
		self.language = language
		
	def checkGame(self):
		languageSupport = support(game=self.game, language=self.language); # Checks language support, if string one or both langauges isn't supported
		if self.game not in support():
			print('Game not found, please check supported games!\n');
			support(True) # prints supported games, in all of their formats
			return False
		elif languageSupport and type(languageSupport) != str:
			return True;
		else:
			print('Language ' + languageSupport + ' not supported in game selected!');
			
	def gamePath(self):
		return support(game=self.game);
			
def core(game, language, type, file, format, audio, create=False):
	global sheet_data;
	global audio_files;
	global csv_data;
	global xlsx_data;
	files = file;
	game_class = games(game=game, language=language)
	# This path is currently for testing purposes
	# Look for game folder to create creationkit settings .txt file, or edit existing .txt
	if create == True:
		drive = os.path.splitdrive(os.getcwd())[0] # Assumes the drive
		print('Checking ' + drive + '\Program Files (x86)\Steam\SteamApps\common \n');
		if os.path.exists(drive + '\Program Files (x86)\Steam\SteamApps\common'):
			print('File found!', end='\r');
			gameClass = games(game=game, language=language);
			mainGame = gameClass.gamePath() # Get the games 'real' name
			special = False # Used to handle unique items (If there's more than one of the same), i.e, 'Skyrim', 'Skyrim Special Edition'
			gamePaths = {'Skyrim': 'Skyrim', 'Skyrim SE': 'Skyrim Special Edition', 'Fallout 4': 'Fallout 4'} # Keys are the filepath after SteamApps\common\
			c_gamepath = gamePaths[mainGame];
			
			if mainGame == 'Skyrim': # Since there's two versions of Skyrim, this checks based on the users previous input
				if game == 'skyrim se' or 'skyrim special edition':
					c_gamepath = gamePaths['Skyrim SE'];
					special = True;
				
			main_path = drive + '\Program Files (x86)\Steam\SteamApps\common\\' + c_gamepath;
			print('Checking ' + main_path + ' for ' + c_gamepath, end='\r\n');
			if os.path.exists(main_path):
				# Path - Skyrim SE 
				print('Found game folder!', end='\r');
				# Check for 'CreationKitCustom.ini' if special is True (This is due to the fact that Skyrim SE makes you create the file, instead of having it already like in Skyrim (vanilla)
				if not os.path.exists(main_path + '/CreationKitCustom.ini'): 
					print('Creating CreationKitCustom.ini...', end='\r');
					with open(main_path + '/CreationKitCustom.ini', "a", encoding='utf-8') as ini: # Creates the ini file with the selected language
						ini.write('[General]\nsLanguage=%s' % gameClass.language);
						
					print('Done!', end='\r');
				else:
					choice_input = input('\rCreationKitCustom.ini already exists! Rewrite? (Y/N)\n');
					if choice_input.lower() == 'y':
						with open(main_path + '/CreationKitCustom.ini', "w", encoding='utf-8') as ini:
							ini.write('[General]\nsLanguage=%s' % gameClass.language);
					else:
						return True;
						
				return False;
			else:
				print('Cannot find path at %s! Are you sure you entered the correct game?' % main_path);
				# add some way to input here 
				
				
			
	# Create the dialogues/ folder
	# Confirm if it exists already or not
	if not os.path.exists('dialogues/'):
		 os.makedirs('dialogues')
			
	filename, fileExt = os.path.splitext(files[0])
	
	# Convert files and place them inside of dialogue folder
	output_files = [];
	
	removeXLSX = True;
	for i in range(0, len(files)):
		# Take the file from its destination if it is not in dialogues/ and put it inside of dialogues/

		g_c = games(game=game, language=language[i])
		if fileExt != '.xlsx':
			output_files.append(convert2(files[i], g_c))
		elif fileExt == '.xlsx':
			output_files.append(files[i]);
			removeXLSX = False;
			
		if fileExt == '.csv':
			convertFromCSV(output_files[i], g_c.gamePath(), language[i]);
		elif fileExt == '.txt' or fileExt == '.xlsx':
			convertFromXLSX(output_files[i], g_c.gamePath(), i, language[i]);
			
	# Remove the leftover XLSX file(s)
	# Confirm output_files are xlsx files
	if removeXLSX == True:
		for checkFiles in output_files:
			f_name, ext = os.path.splitext(checkFiles)
			if ext!= '.xlsx':
				break;
			else:
				os.remove(checkFiles);

	file_name = game_class.gamePath() + '_Dialogue_' + '_'.join(language)
	mainData = ''
	
	if fileExt == '.csv':
		mainData = csv_data;
	elif fileExt == '.xlsx' or fileExt == '.txt':
		mainData = xlsx_data;
		
		
	c_data = [];
	count = 0;

	for x in mainData: # language 1, language 2
		for i in x.keys():
			# Find name of key in other dict
			try:
				c_data.append([mainData[0][i], mainData[1][i]])
			except KeyError:
				pass
			count += 1;
		sheet_data = c_data;
		break;
			
	# remove duplicates
	def removeDuplicates(sheet, language):
		data = [];
		for i in sheet:
			if i not in data:
				data.append(i);
		data[0][0] = 'RESPONSE TEXT-' + language[0];
		data[0][1] = 'RESPONSE TEXT-' + language[1];
		return data;
	
	# if format = txt
	if format.lower() == 'txt':
		with open(file_name + '.txt', 'w', encoding='utf-8') as prc:
			for i in removeDuplicates(sheet_data, language):
				try:
					if i != ' ' or i != None:
						prc.write(';'.join(i) + '\n')
				except:
					pass
				
	# if format = csv
	if format.lower() == 'csv':
		with open(file_name + '.csv', 'w', encoding='utf-8') as csv_export:
			csv_writer = csv.writer(csv_export, delimiter='\t');
			# Find [Sound:, and ending ] indexes
			a_count = 0;
			for i in removeDuplicates(sheet_data, language):
				def removeAudio(string_of):
					try:
						x = string_of.index(']');
						return string_of[x + 2:]
					except ValueError:
						return string_of;
				for rows in range(0, len(i)):
					i[rows] = removeAudio(i[rows]);
					
				i.append(audio_files[a_count].lower())
				csv_writer.writerow(i)
				a_count += 1;

	
	# If audio is not False
	if audio:
		print('Looking for ' + str(audio[0]) + ' file in audio/...');
		
		# Create the audio/ folder
		# Confirm if it exists already or not
		if not os.path.exists('audio/'):
			os.makedirs('audio')
			
		# Look for a file named what was passed to the --audio argument
		folders = os.listdir('audio/');
		main_folder = '';
		for find_folder in folders:
			if find_folder.lower() == audio[0].lower():
				print('Found folder ' + str(find_folder) + '!');
				main_folder = find_folder;
				break;
				
		if main_folder != '':
			a_arr = [];
			# Locate all audio file(s) within audio_files array 
			# Loop through all folders and add their contents to a new array
			# Example: Folder_1 > Folder_2 > Folder_3 > Audio_Folder > audio_file.fuz
			for root, dirs, files in os.walk('audio\\' + main_folder):
				# paths to array
				for file in files:
					if file in audio_files:
						a_arr.append(os.path.join(root, file));
						
			# Create folder
			new_folder = 'audio/' + game_class.gamePath() + '_' + audio[0] + '_audio'
			if not os.path.exists(new_folder): 
				os.makedirs(new_folder)
				
			# Abbreviated language
			abr = audio[0][:2].upper()
			for audioFile in a_arr:
				# Move and rename audioFile i.e, voice_01 > voice_01_EN
				# Move to folder named GAMENAME_LANGUAGE_AUDIO within the audio folder, be sure file does not already exist
				# ???
				# profit
				audioF = audioFile.split('\\');
				audioF = audioF[len(audioF) - 1]
				audioF = audioF.split('.')[0] # The files name without .fuz
				try:
					os.rename(audioFile, new_folder + '/' + audioF + '_' + abr + '.fuz');
				except FileExistsError:
					pass;
		else:
			print('Couldn\'t find folder ' + audio[0] + '..');
			
def main():
	global supportedFormats;
	parser = argparse.ArgumentParser(description=m_descr);
	#python core.py -g skyrim -l english spanish -t txt -f dialogues/export1.txt dialogues/export2.txt
	#python core.py -g skyrim -l english spanish -t txt -f dialogues/export1.txt dialogues/export2.txt -a russian
	
	parser.add_argument('-g', '--game', nargs='+', 
		help="Source game to process")
	
	parser.add_argument('-l', '--language', nargs='+',  
		help="Language to process")
		
	parser.add_argument('-t', '--type',  
		help="Type of format. Supported formats: [txt, csv]")
		
	parser.add_argument('-f', '--file', nargs='+',
		help='The export file')
		
	# Audio isn't grabbed by default, arg is optional
	parser.add_argument('-a', '--audio', nargs='+',
		help='Grab audio *optional')
		
	# FOR TESTING PURPOSES; python core.py -c -g YOUR_GAME_NAME** 
	# parser.add_argument('-c', '--create', action='store_true',
	#	default=False,
	#	help='Creates or edits .ini file for selected game')
	
	args = parser.parse_args()
	
	# See if user stated two files, if not pass, and pass args.file false
	try:
		if len(args.file) < 2:
			print('Note: Only %s file(s) stated!' % len(args.file));
	except:
		args.file = False
		pass
	
	if len(args.game) > 1:
		args.game = ' '.join(args.game); # i.e, ['fallout', '4'] => 'fallout 4'
	else:
		args.game = args.game[0]
	
	if len(args.language) < 2:
		print('Note: Only %s language(s) stated! Please use more than one language, (i.e English Spanish)' % len(args.file));
		return False
	
	for c in range(0, len(args.language)):
		args.language[c] = args.language[c][:1].upper() + args.language[c][1:].lower() # from 'spanish', to 'Spanish'
	
	if not args.type:
		args.type = 'txt'
	elif args.type.lower() not in supportedFormats:
		print(args.type + ' format is not supported! Please see the supported formats: ' + str(supportedFormats));
		return False;
		
	# If audio is empty, pass it False
	if args.audio == None:
		args.audio = False;

	gameClass = games(args.game.lower(), '', args.language);
	
	if gameClass.checkGame(): # Checks to see if 'game' is in the supported games arr
		core(game = args.game.lower(),
		language = args.language,
		type = args.type,
		file = args.file,
		format = args.type,
		audio = args.audio)
		#create = args.create)
	
if __name__ == '__main__':
	main()